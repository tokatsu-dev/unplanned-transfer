import os
import pandas as pd
import openpyxl
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta
import math
import jaconv
import re
import openpyxl
import os
import pandas as pd
import copy
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

import pandas as pd
import re


import pandas as pd

import pandas as pd
import openpyxl
import csv



import pandas as pd
import openpyxl

import pandas as pd

import pandas as pd
import openpyxl
from openpyxl import Workbook
import xlrd
import csv


import os
import glob
import pandas as pd
from typing import List, Tuple, Optional, Union, Any, Dict
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet


import os
import glob
import pandas as pd
from IPython.display import display



def load_target_iraisho_files(mapping, reference_folder_path, extensions):
    """
    指定された mapping 情報に基づき、書類パターンごとに、配送依頼書ファイルを検索・読み込み・加工する。

    Parameters
    ----------
    mapping : dict
        書類名・ヘッダー行などを含むマッピング情報。
    reference_folder_path : str
        ファイル探索対象のフォルダパス。
    extensions : list[str]
        検索対象の拡張子リスト（例: ["xlsx", "xlsm"]）。
    is_number : function
        商品コードが数値かどうかを判定する関数。
    load_excel_like : function
        Excelファイルを読み込み、(DataFrame, Workbook, Worksheet) を返す関数。

    Returns
    -------
    list[tuple[pd.DataFrame, Workbook, Worksheet]]
        各ファイルについて処理後の DataFrame と対応する Workbook, Worksheet のタプルを格納したリスト。
    """

    file_name = mapping['書類名']
    target_files = []

    # --- ファイル検索 ---
    for ext in extensions:
        file_pattern = os.path.join(reference_folder_path, f"*{file_name}*.{ext}")
        matched_files = glob.glob(file_pattern)
        target_files.extend(matched_files)

    print(f"検索結果: {target_files}")

    if not target_files:
        print(f"⚠ ファイルが見つかりません: {file_name}")
        return []

    skip_rows = int(mapping.get("ヘッダー行", 0))
    processed_results = []

    # --- 各ファイル処理 ---
    for target_file in target_files:
        basename = os.path.splitext(os.path.basename(target_file))[0]
    
        target_df, target_ws = load_excel_like(target_file, skip_rows)

        # --- 商品コード列での有効行抽出 ---
        item_code_col = mapping.get("品番")
        if item_code_col not in target_df.columns:
            # 現行仕様のヨシケイ確定手配書のための特別処理。削除対象
            if item_code_col == "商品コード":
                hinban = "商品ｺｰﾄﾞ"
                if hinban not in target_df.columns:
                    print(f"⚠ '{item_code_col}'列も'{hinban}'列も参照中の書類に存在しません → {target_file}")
                    continue
                else:
                    print(f"⚠ '{item_code_col}'列が存在しないため、'{hinban}'列を代わりに使用します → {target_file}")  

        valid_mask = target_df[item_code_col].apply(is_number)

        if valid_mask.any():
            last_row = valid_mask[valid_mask].index[-1]
            target_df = target_df.loc[:last_row]
        else:
            target_df = pd.DataFrame(columns=target_df.columns)

        # --- 空の一時DFを作成 ---
        create_temp_df = pd.DataFrame(index=range(len(target_df)))

        processed_results.append((basename, target_df, target_ws, create_temp_df))

    return processed_results



def load_target_hokokusho_files(mapping, reference_folder_path, extensions):
    """
    書類名に対応するファイルを検索して読み込み、
    DataFrame, Workbook, Worksheet, 空の一時DataFrameを返すリストを作成。

    Parameters
    ----------
    mapping : dict
        '書類名' と 'ヘッダー行' が含まれる辞書
    reference_folder_path : str
        ファイル検索の基準フォルダ
    extensions : list of str
        検索するファイル拡張子リスト（例: ["xlsx", "xls", "csv"]）

    Returns
    -------
    results : list of tuples
        各要素は (target_df, target_ws, create_temp_df)
    """

    file_name = mapping['書類名']
    target_files = []

    # --- ファイル検索 ---
    for ext in extensions:
        file_pattern = os.path.join(reference_folder_path, f"*{file_name}*.{ext}")
        matched_files = glob.glob(file_pattern)
        target_files.extend(matched_files)

    print(f"検索結果: {target_files}")

    if not target_files:
        print(f"⚠ ファイルが見つかりません: {file_name}")
        return []

    skip_rows = int(mapping.get("ヘッダー行", 0))
    results = []

    # --- 各ファイル処理 ---
    for target_file in target_files:
        basename = os.path.splitext(os.path.basename(target_file))[0]
        target_df, target_ws = load_excel_like(target_file, skip_rows)
        print(f"{len(target_df)} 行を読み込みました: {os.path.basename(target_file)}")

        # --- 空の一時DFを作成 ---
        create_temp_df = pd.DataFrame(index=range(len(target_df)))

        results.append((basename, target_df, target_ws, create_temp_df))

    return results



def load_excel_like(target_file: str, skip_rows: int) -> Tuple[pd.DataFrame, OpenpyxlWorkbook, OpenpyxlWorksheet]:
    """
    ExcelやCSVファイルを読み込み、
    DataFrame, Workbook, Worksheet の3つを返す。
    """
    ext = target_file.lower().split('.')[-1]

    # === xlsx ===
    if ext == "xlsx":
        df = pd.read_excel(target_file, skiprows=skip_rows, engine="openpyxl")
        wb = openpyxl.load_workbook(target_file, data_only=True)
        ws = wb.active
        return df, ws

    # === xls ===
    elif ext == "xls":
        try:
            import xlrd
            df = pd.read_excel(target_file, skiprows=skip_rows, engine="xlrd")
            wb, ws = dataframe_to_workbook(df)
            return df, ws
        except Exception as e1:
            print(f"xls読込失敗({type(e1).__name__})、CSVとして再試行します。")
            try:
                df = load_csv_safe(target_file, skip_rows)
                wb, ws = dataframe_to_workbook(df)
                return df, ws
            except Exception as e2:
                raise ValueError(
                    f"xls拡張子のファイルですが、Excel/CSVいずれでも開けませんでした: {target_file}\n"
                    f"原因: {type(e1).__name__}, {type(e2).__name__}"
                )

    # === csv ===
    elif ext == "csv":
        df = load_csv_safe(target_file, skip_rows)
        wb, ws = dataframe_to_workbook(df)
        return df, ws

    else:
        raise ValueError(f"未対応の拡張子: {ext}")



def load_csv_safe(path: str, skip_rows: int = 0) -> pd.DataFrame:
    """
    CSV/TSVファイルを確実に読み込む。
    区切り文字やエンコーディングを自動判定。
    """
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "euc_jp"]

    for enc in encodings:
        try:
            # .xlsファイルはだいたいタブ区切りCSVである。区切り文字を判定するため、最初に少しだけ読む
            with open(path, "r", encoding=enc, errors="ignore") as f:
                sample = f.read(2048)

            # 区切り文字を自動判定
            if "\t" in sample:
                sep = "\t"
            elif ";" in sample:
                sep = ";"
            else:
                sep = ","

            # 読み込み実行
            df = pd.read_csv(path, encoding=enc, skiprows=skip_rows, sep=sep, engine="python")

            # --- カラム名がタブ混入している場合の補正 ---
            df.columns = [c.replace("\t", "").strip() for c in df.columns]

            print(f"読み込み成功: {enc}, 区切り='{sep}', shape={df.shape}")
            return df

        except Exception as e:
            print(f"失敗: {enc} -> {type(e).__name__}")

    raise ValueError(f"すべてのエンコーディングで読み込み失敗: {path}")






def dataframe_to_workbook(df: pd.DataFrame) -> Tuple[OpenpyxlWorkbook, OpenpyxlWorksheet]:
    """
    DataFrameをopenpyxlのWorkbookオブジェクトに変換する。
    CSVでもxls偽装ファイルでもセル参照を可能にするため。
    """
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return wb, ws


def calc_formula(df: pd.DataFrame, val: str) -> pd.Series:
    """
    df: 対象のDataFrame
    val: 計算式文字列または単一列名
         例:
         - "入数×出荷数1" → 計算
         - "ロットNo" → そのまま列を返す
    """
    # print(f"元の入力: {val}")

    # 演算子を含むかチェック（全角も含む）
    if re.search(r"[×÷＋−\+\-\*/]", val):
        # 演算記号の置換（全角→半角）
        replacements = {"×": "*", "÷": "/", "＋": "+", "−": "-"}
        expr = val
        for k, v in replacements.items():
            expr = expr.replace(k, v)
        # print(f"変換後の演算式: {expr}")

        # トークン抽出
        tokens = re.split(r"[\+\-\*/]", expr)
        tokens = [t.strip() for t in tokens if t.strip()]
        # print(f"演算に使用するカラム: {tokens}")

        # 演算式をベクトル指向で置換
        expr_vec = expr
        for token in tokens:
            if token in df.columns:
                # Series を丸ごと置換
                expr_vec = expr_vec.replace(token, f"df['{token}']")
            else:
                expr_vec = expr_vec.replace(token, "0")

        # print(f"eval前の式: {expr_vec}")
        result = eval(expr_vec)
        return result

    else:
        # 単一カラム名の場合はそのまま Series を返す
        if val in df.columns:
            # print(f"単一列参照として処理: {val}")
            return df[val]
        else:
            print(f"指定列が存在しません: {val}")
            return pd.Series([None]*len(df))



def split_kikaku_series(df: pd.DataFrame) -> Tuple[pd.Series, pd.Series, pd.Series]:
    """
    規格列を分割して 4桁ロット、入数、合 の Series を返す関数。

    args:
        df: 報告書の生データフレーム（規格列を含む）

    returns:
        lot_series: 4桁ロット列の Series
        count_series: 入数列の Series
        total_series: 合列の Series
    """
    lot_list = []
    count_list = []
    total_list = []

    for i, row in df.iterrows():
        kikaku = row.get("規格", None)

        # 初期値
        lot_val = None
        count_val = None
        total_val = "1"  # デフォルトは1

        if pd.notna(kikaku):
            # 【】内の4桁ロット
            match = re.search(r"【([^】]+)】", str(kikaku))
            if match:
                lot_part = match.group(1)
                lot_part = jaconv.z2h(lot_part, kana=False, digit=True, ascii=True)
                if len(lot_part) >= 4 and lot_part.isdigit():
                    lot_val = lot_part[-4:]

            # "×"で分割して入数・合を取得
            parts = re.split(r"[×xX]", str(kikaku))
            if len(parts) >= 2:
                count_val = parts[1].strip()
            if len(parts) >= 3 and "合" in parts[2]:
                total_val = parts[2].replace("合", "").strip()

        lot_list.append(lot_val)
        count_list.append(count_val)
        total_list.append(total_val)

    lot_series = pd.Series(lot_list, index=df.index)
    count_series = pd.Series(count_list, index=df.index)
    total_series = pd.Series(total_list, index=df.index)

    return count_series, total_series, lot_series



def get_cell_value_by_cell_reference(input_ws: Any, cell_ref: str) -> str:
    """
    セル番号で値を取得する関数

    Parameters
    ----------
    input_ws : xlwings.Sheet
        値を取得する対象のワークシート
    cell_ref : str
        セル参照（例: 'A1', 'C10' など）

    Returns
    -------
    str
        セルの値（空セルの場合は空文字）
    """
    # print(f"セル参照の取得を試みます: {cell_ref}")

    if re.match(r"^[A-Za-z]+\d+$", cell_ref):
        # print(f"セル参照として認識: {cell_ref}")
        cell_value = input_ws[cell_ref].value
        if cell_value is None:
            return ""
        else:
            # print(f"依頼書の{cell_ref}セルから検出：{cell_value}")
            return str(cell_value).strip()
    else:
        # 参照文字列でない場合
        return cell_ref


def get_cell_value_by_column_reference(i: int, target_df: pd.DataFrame, val: str) -> Optional[str]:
    """
    指定されたDataFrame列（val）から、i行目の値を取得して返す関数。
    key: ログ出力用の項目名
    val: 列名
    """

    if val in target_df.columns:
        cell_value = target_df[val].iloc[i]

        # float型なら整数に変換（例: 20260301.0 → 20260301）
        if isinstance(cell_value, float) and cell_value.is_integer():
            cell_value = int(cell_value)

        # 文字列化して前後の空白を除去
        cell_value = str(cell_value).strip()

        # print(f"{key}キーを依頼書データフレームの列 '{val}' から検出：{cell_value}")
        return cell_value

    else:
        # print(f"{key}キーに対応する依頼書データフレームの列 '{val}' が存在しません。")
        return None



def process_date_value(val: Union[int, float, datetime, pd.Timestamp, str, pd.Series], type: int) -> Union[pd.Series, Optional[str]]:
    """
    日付形式を統一する関数。Seriesが渡された場合は再帰的に処理する。

    Args:
        val: 日付を表す値（int, float, datetime, pd.Timestamp, str, または pd.Series）
        type: 1: 日付出力 / 2: ロットNo出力（4桁 or YYYYMMDD形式）
    Returns:
        YYYYMMDD形式の日付 or 4桁ロット番号
    """

    # Seriesが渡された場合は apply で再帰呼び出し
    if isinstance(val, pd.Series):
        return val.apply(lambda x: process_date_value(x, type))

    fmt_out = "%Y/%m/%d" if type == 1 else "%Y%m%d"

    # 数値の場合（整数・浮動小数点）
    if isinstance(val, (int, float)):
        if 1000 <= int(val) <= 9999:  # 4桁ロット
            return str(int(val))
        try:  # Excelシリアル値として日付変換
            date = datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(val))
            return date.strftime(fmt_out)
        except Exception:
            return None

    # datetime 型
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.strftime(fmt_out)

    # 文字列の場合
    if isinstance(val, str):
        s = val.strip()
        s = s.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        s = s.replace("年", "/").replace("月", "/").replace("日", "")
        s = re.sub(r"（.*?）|\(.*?\)", "", s).strip()


        # 文字列が4桁ロットかどうか判定
        if s.isdigit() and 1000 <= int(s) <= 9999:
            return s

        # 日付フォーマットを順に試す
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S",
                    "%Y/%m/%d", "%Y/%m/%d %H:%M:%S",
                    "%Y%m%d", "%Y.%m.%d"):
            try:
                parsed = datetime.strptime(s, fmt)
                return parsed.strftime(fmt_out)
            except ValueError:
                continue
        return None

    # その他の型は None
    return None





def make_lot_no(lot_4digits: Any, exp_date: Optional[Union[datetime, pd.Timestamp, str]] = None) -> Optional[str]:
    """
    4桁ロットと賞味期限（exp_date）から最終的なロットNoを作成する関数。

    Args:
        lot_4digits: 4桁ロット番号（文字列または数字）
        exp_date: 賞味期限（datetime, pd.Timestamp, または文字列形式）

    Returns:
        str: 結合済みロットNo（YYYYMMDD-xxxx またはYYYYMMDD）
    """

    # 4桁ロットがない場合
    if lot_4digits is None or pd.isna(lot_4digits):
        # 賞味期限がなければNoneを返す
        if exp_date is None:
            return None
        
        # 賞味期限があればYYYYMMDD形式に
        if isinstance(exp_date, (datetime, pd.Timestamp)):
            date_str = exp_date.strftime("%Y%m%d")
        else:
            date_str = re.sub(r"\D", "", str(exp_date))
        return date_str

    # 4桁ロットがある場合
    lot_str = str(lot_4digits).strip()

    # 賞味期限もある場合、YYYYMMDD-XXXX形式に
    if exp_date is not None:
        if isinstance(exp_date, (datetime, pd.Timestamp)):
            date_str = exp_date.strftime("%Y%m%d")
        else:
            date_str = re.sub(r"\D", "", str(exp_date))
        lot_no = f"{date_str}-{lot_str}"
    
    # 賞味期限がなければNoneを返す
    else:
        lot_no = None

    if lot_no is not None:
        # 全角→半角に変換（数字・ハイフン）
        lot_no = lot_no.translate(str.maketrans({
            "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
            "５": "5", "６": "6", "７": "7", "８": "8", "９": "9"
        }))


    return lot_no




# 報告書データフレームと依頼書で製品の突合を行う関数
def fill_lot_No(
    i: int,
    input_ws: Any,
    start_row: int,
    folder_or_filename: str,
    vals_for_lot_no: Dict[str, Any],
    report_df_narrowed: pd.DataFrame,
) -> Union[str, pd.Series]:
    # 突合に必要な情報：品番、賞味期限、総パック数
    # ロットNo作成に必要な情報：賞味期限、規格

    print("報告書サマリのカラム一覧：", report_df_narrowed.columns)

    # 型の不一致による突合失敗を防ぐため文字列化
    report_df_narrowed["品番"] = report_df_narrowed["品番"].astype(str)
    report_df_narrowed["移送数"] = report_df_narrowed["移送数"].astype(str)


    
    # 1. 突合に必要な依頼書からの情報を整理する

    # 突合条件の一覧、Noneのキーは除外
    dict_from_iraisho = {
        k: v
        for k, v in {
            "品番": vals_for_lot_no["品番"],
            # "賞味期限": vals_for_lot_no["賞味期限"],
            # "渡し先名": vals_for_lot_no["渡し先名"],
            "移送数": vals_for_lot_no["移送数"]
        }.items()
        if v is not None and not pd.isna(v) 
    }

    print(f"次の依頼書データで突合を行います: {dict_from_iraisho}")



    # 2. 報告書データフレームからの商品特定
    cond = pd.Series(True, index=report_df_narrowed.index)
    for key, val in dict_from_iraisho.items():
        if key not in report_df_narrowed.columns:
            print(f"⚠️報告書データに {key} 列がありません。スキップします。")
            continue

        # 文字列に揃えて比較（余計な空白も削除）
        cond = cond & (report_df_narrowed[key].astype(str).str.strip() == str(val).strip())
    matched = report_df_narrowed[cond] # ヒットした行を格納。列は特にカットされない


    # 【cond &= ...の問題について】

    #    品番
    # 0  A123
    # 1  B456
    # 2  C789

    # ここでreport_df_narrowed["品番"] == "A123"を評価すると

    # 0     True
    # 1    False
    # 2    False
    # Name: 品番, dtype: bool

    # のようなpandas.Series（長さ3の真偽リスト）が返ってくる。ここで、１回目のループ
    # cond &= (report_df_narrowed["品番"] == "A123")    をpythonは
    # cond = cond & (report_df_narrowed["品番"] == "A123")  と解釈する。「左側の cond = True」と「右側の Series」でブール演算をしようとする。当然比較できないので、Series 全体を True として扱うことがある。結果的にcondはSeriesではなく１個のtrueになる。すると２回目のループでは
    # cond &= (report_df_narrowed["賞味期限"] == "20251010")   このときcondは True なので、pandas Series との比較が成立せず、また同じことが繰り返される。最終的に
    # matched = report_df_narrowed[cond]    で cond は True なので、report_df_narrowed 全体が matched に入ってしまう。
    # だから    if matched.empty:   に流れない。


    print(f"ロットNo構築結果: \n{matched}")

    if matched.empty:
        print("該当商品は報告書データに存在しません")
        return "報告書データに該当商品なし"
    else:
        return matched["ロットNo"] # matched DataFrame の "ロットNo" 列だけを抜き出した pandas.Series
        # 該当行すべてのデータフレームが返ってくる。つまり品番がヒットしたものはすべて転記される





# 賞味期限の形式を統一する関数
def unify_date_format(lot_no: Any, exp_date_lot_no: Optional[Any] = None) -> Optional[str]:

    print(f"賞味期限の形式を統一します。ロットNo:{lot_no}, 賞味期限:{exp_date_lot_no}")

    if lot_no is None or pd.isna(lot_no):
        if exp_date_lot_no is not None:
            # 4桁ロットがNoneでもexp_date_lot_noがあればそれでロットNo構築
            if isinstance(exp_date_lot_no, (datetime, pd.Timestamp)):
                exp_date_str = exp_date_lot_no.strftime("%Y%m%d")

            # さらにこの場合でexp_date_lot_noが数字５桁ならExcelシリアル値として変換
            elif isinstance(exp_date_lot_no, (int, float)) and 10000 <= exp_date_lot_no < 100000:
                dt = datetime(1899, 12, 30) + timedelta(days=exp_date_lot_no)
                exp_date_str = dt.strftime("%Y%m%d")

            else:
                exp_date_str = str(exp_date_lot_no).strip()
                # もし "2026-08-31 00:00:00" のような形式なら数字だけ抜き出し
                exp_date_str = re.sub(r"\D", "", exp_date_str)
            return exp_date_str
        else:
            return None
    

    # ハイフンならスキップ
    if str(lot_no).strip() == "-":
        return lot_no

    # datetime型ならそのまま整形
    if isinstance(lot_no, (datetime, pd.Timestamp)):
        return lot_no.strftime("%Y%m%d")

    # 数字５桁ならExcelシリアル値と認識して変換
    if isinstance(lot_no, (int, float)) and 10000 <= lot_no < 100000:
        # Excelのシリアル値 → datetime
        dt = datetime(1899, 12, 30) + timedelta(days=lot_no)

        # YYYY-MM-DD HH:MM:SS形式になるのでYYYYMMDD形式に変換して返す
        return dt.strftime("%Y%m%d")



    # 以降、加工のため文字列として扱う
    s = str(lot_no).strip()
    # 小数点付き数字は整数に変換（例: 20260301.0 → 20260301）
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]


    # 4桁ロットの場合、依頼書の賞味期限からロットNoを結合して返す。なのでこの場合報告書データフレームを参照することはない
    if re.fullmatch(r"\d{4}", s):
        if exp_date_lot_no is not None:
            # 日付型であることが多いので文字列に整形
            if isinstance(exp_date_lot_no, (datetime, pd.Timestamp)):
                exp_date_str = exp_date_lot_no.strftime("%Y%m%d")
            else:
                exp_date_str = str(exp_date_lot_no).strip()
                # もし "2026-08-31 00:00:00" のような形式なら数字だけ抜き出し
                exp_date_str = re.sub(r"\D", "", exp_date_str)

            print(f"賞味期限{exp_date_str}の整形を行いました。")
            
            combined = f"{exp_date_str}-{s}"
            # 全角→半角に変換（数字・ハイフン）
            combined = combined.translate(str.maketrans({
                "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
                "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
                "－": "-"
            }))
            return combined
        else:
            return s  # 賞味期限が指定されていない場合
        

    try:
        if "年" in s:
            dt = datetime.strptime(s, "%Y年%m月%d日")
            formatted = dt.strftime("%Y%m%d")
        elif "." in s:
            dt = datetime.strptime(s, "%Y.%m.%d")
            formatted = dt.strftime("%Y%m%d")
        elif re.fullmatch(r"\d{8}", s):  # すでに8桁数字
            formatted = s
        # else:
        #     # 数字以外を除去してみる（保険）
        #     formatted = re.sub(r"\D", "", s)
        #     if len(formatted) == 8:
        #         pass
        #     else:
        #         print(f"⚠️ 不明な形式: {s} -> {formatted}")
        return formatted

    except Exception as e:
        print(f"⚠️ パース失敗: {s} ({e})")
        return s
    

    

def add_prefix(x: Any) -> Any:
    if pd.isna(x):
        return x
    s = str(x).strip()
    # 数字だけ & 桁数チェック
    if re.fullmatch(r"\d{5}", s):
        return "360" + s
    elif re.fullmatch(r"\d{6}", s):
        return "36" + s
    else:
        return s
    
    # valid_maskの操作と一部重複する？




# 数字として解釈できる行だけTrueにするマスク
def is_number(x: Any) -> Union[bool, pd.Series]:
    if isinstance(x, pd.Series):
        return x.apply(is_number)

    # pd.isnaはstr型を検出できないため、
    if pd.isna(x):
        return False

    x = str(x)
    try:
        float(x)
        return True
    except (ValueError, TypeError):
        return False


#データベースに接続
import pandas as pd
import pymssql
# ==============================
# SQL Serverからデータ取得
# ==============================
def fetch_sql_df(server, database, username, password, sql_query):
    """
    server(str): SQL Serverのサーバー名またはIPアドレス
    database(str): データベース名
    username(str): ユーザー名
    password(str): パスワード
    sql_query(str): 実行するSQLクエリ
    SQL Serverへ接続し、指定したSQLクエリを実行して結果をDataFrameで返す。
    接続は都度開いて、取得後に閉じる。
    """
    # DB接続
    conn = pymssql.connect(
        server=server,
        user=username,
        password=password,
        database=database,
        charset="UTF-8",
    )
    # SQL実行→DataFrame化
    df = pd.read_sql(sql_query, conn)
    # コネクションを閉じる
    conn.close()
    return df